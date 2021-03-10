import requests
import openpyxl
import warnings
import pyproj
import sys
from flask import Flask

warnings.filterwarnings("ignore")

app = Flask(__name__)
wb = openpyxl.load_workbook('2018_01_Sites_mobiles_2G_3G_4G_France_metropolitaine_L93.xlsx')
data_set = []
for i in range(2, wb.active.max_row + 1):
    data_set.append([
        wb.active.cell(i, 1).value,
        wb.active.cell(i, 2).value,
        wb.active.cell(i, 3).value,
        wb.active.cell(i, 4).value,
        wb.active.cell(i, 5).value,
        wb.active.cell(i, 6).value
    ])
wgs84 = pyproj.CRS("EPSG:4326")
lambert = pyproj.CRS("epsg:2154")


def get_coordinates(address):
    # def get_coordinates(city="", postcode="", address=""):
    r = requests.get("https://api-adresse.data.gouv.fr/search/?q=" + str(address))
    # if postcode != "":
    #     r = requests.get("https://api-adresse.data.gouv.fr/search/?q=8+bd+du+port&postcode=" + str(postcode).lower())
    # elif city != "":
    #     r = requests.get('https://api-adresse.data.gouv.fr/search/?q=' + str(city).lower() + '&type=street')
    # else:
    #     r = requests.get("https://api-adresse.data.gouv.fr/search/?q=" + str(address))
    try:
        return r.json()['features'][0]['geometry']['coordinates']
    except:
        print(r.json())
        print("Unexpected error:", sys.exc_info()[0])
        raise


@app.route('/<address>', methods=['GET'])
def index(address):
    x, y = get_coordinates(address)
    lambert_x, lambert_y = pyproj.transform(wgs84, lambert, y, x)

    output_data_set_tmp = []
    for data in data_set:
        if data[1] != '#N/A' and data[2] != '#N/A':
            if abs(int(data[1]) - int(lambert_x)) < 5000 and abs(int(data[2]) - int(lambert_y)) < 5000:
                output_data_set_tmp.append([data[0], data[3], data[4], data[5]])

    output_data_set_clean = []
    for data in output_data_set_tmp:
        flg = True
        for data_for_check in output_data_set_clean:
            if data_for_check[0] == data[0]:
                if data[1] == 0 and data[1] != data_for_check[1]:
                    data[1] = 1
                if data[2] == 0 and data[2] != data_for_check[2]:
                    data[2] = 1
                if data[3] == 0 and data[3] != data_for_check[3]:
                    data[3] = 1
                if data not in output_data_set_clean:
                    output_data_set_clean.remove(data_for_check)
                    output_data_set_clean.append(data)
                    flg = False
        if data not in output_data_set_clean and flg:
            output_data_set_clean.append(data)

    wb2 = openpyxl.load_workbook('operators.xlsx')
    operators = []
    for i in range(1, wb2.active.max_row + 1):
        operators.append([
            wb2.active.cell(i, 1).value,
            wb2.active.cell(i, 2).value
        ])
        
    output_data_set_translated = []
    for data in output_data_set_clean:
        for operator in operators:
            if operator[0] == data[0]:
                data[0] = operator[1]
                if data[1] == 1:
                    data[1] = 'True'
                else:
                    data[1] = 'False'
                if data[2] == 1:
                    data[2] = 'True'
                else:
                    data[2] = 'False'
                if data[3] == 1:
                    data[3] = 'True'
                else:
                    data[3] = 'False'
                output_data_set_translated.append(data)

    response = '{'
    for data in output_data_set_translated:
        response = response + "\"" + str(data[0]) + "\": {\"2G\": " + str(data[1]) + \
                   ", \"3G\": " + str(data[2]) + ", \"4G\": " + str(data[3]) + "}, "
    response = response[:-2] + '}'
    # return json.dumps(output_data_set_translated)
    return response


if __name__ == "__main__":
    app.run(debug=False)
